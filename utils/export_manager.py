from typing import Dict, List, Optional
import csv
import json
from datetime import datetime
from PyQt6.QtWidgets import QFileDialog, QMessageBox, QWidget
from PyQt6.QtCore import QObject, pyqtSignal, QThread
from PyQt6.QtGui import QPixmap, QPainter
import matplotlib.pyplot as plt
import matplotlib.dates as mdates
from matplotlib.backends.backend_agg import FigureCanvasAgg
import pandas as pd
import os

try:
    from reportlab.lib.pagesizes import A4
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
    from reportlab.lib.units import cm
    from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle, Image
    from reportlab.lib import colors
    from reportlab.lib.enums import TA_CENTER, TA_LEFT, TA_RIGHT
    REPORTLAB_AVAILABLE = True
except ImportError:
    REPORTLAB_AVAILABLE = False

class ExportThread(QThread):
    """Thread für Export-Operationen"""
    
    progress_updated = pyqtSignal(int)
    export_finished = pyqtSignal(bool, str)  # success, message
    
    def __init__(self, export_type: str, file_path: str, data: Dict):
        super().__init__()
        self.export_type = export_type
        self.file_path = file_path
        self.data = data
    
    def run(self):
        """Führt den Export durch"""
        try:
            if self.export_type == 'pdf':
                self._export_pdf()
            elif self.export_type == 'csv':
                self._export_csv()
            elif self.export_type == 'excel':
                self._export_excel()
            elif self.export_type == 'json':
                self._export_json()
            
            self.export_finished.emit(True, f"Export erfolgreich: {self.file_path}")
            
        except Exception as e:
            self.export_finished.emit(False, f"Export-Fehler: {str(e)}")
    
    def _export_pdf(self):
        """Exportiert als PDF"""
        if not REPORTLAB_AVAILABLE:
            raise ImportError("ReportLab ist nicht installiert. Bitte installieren Sie es mit: pip install reportlab")
        
        self.progress_updated.emit(10)
        
        doc = SimpleDocTemplate(self.file_path, pagesize=A4)
        styles = getSampleStyleSheet()
        story = []
        
        # Titel
        title_style = ParagraphStyle(
            'CustomTitle',
            parent=styles['Heading1'],
            fontSize=24,
            spaceAfter=30,
            alignment=TA_CENTER
        )
        story.append(Paragraph("BAK-Kalkulator Bericht", title_style))
        story.append(Spacer(1, 20))
        
        # Datum
        date_style = ParagraphStyle(
            'DateStyle',
            parent=styles['Normal'],
            fontSize=12,
            alignment=TA_RIGHT
        )
        story.append(Paragraph(f"Erstellt am: {datetime.now().strftime('%d.%m.%Y %H:%M')}", date_style))
        story.append(Spacer(1, 30))
        
        self.progress_updated.emit(20)
        
        # Personendaten
        if 'person_data' in self.data:
            person = self.data['person_data']
            story.append(Paragraph("Personendaten", styles['Heading2']))
            
            person_table_data = [
                ['Geschlecht:', person.get('gender', 'N/A')],
                ['Alter:', f"{person.get('age', 'N/A')} Jahre"],
                ['Größe:', f"{person.get('height', 'N/A')} cm"],
                ['Gewicht:', f"{person.get('weight', 'N/A')} kg"],
                ['BMI:', f"{person.get('bmi', 'N/A'):.1f}" if person.get('bmi') else 'N/A'],
                ['Körperfettanteil:', f"{person.get('body_fat', 'N/A')}%"],
                ['Trinkgewohnheit:', person.get('drinking_habit', 'N/A')]
            ]
            
            person_table = Table(person_table_data, colWidths=[4*cm, 6*cm])
            person_table.setStyle(TableStyle([
                ('BACKGROUND', (0, 0), (0, -1), colors.lightgrey),
                ('TEXTCOLOR', (0, 0), (-1, -1), colors.black),
                ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
                ('FONTNAME', (0, 0), (-1, -1), 'Helvetica'),
                ('FONTSIZE', (0, 0), (-1, -1), 10),
                ('BOTTOMPADDING', (0, 0), (-1, -1), 6),
                ('BACKGROUND', (1, 0), (1, -1), colors.beige),
            ]))
            story.append(person_table)
            story.append(Spacer(1, 20))
        
        self.progress_updated.emit(40)
        
        # Getränke
        if 'drinks_data' in self.data and self.data['drinks_data']:
            story.append(Paragraph("Konsumierte Getränke", styles['Heading2']))
            
            drinks_table_data = [['Getränk', 'Menge (ml)', 'Alkohol (%)', 'Zeit', 'Alkohol (g)']]
            
            for drink in self.data['drinks_data']:
                alcohol_grams = drink['volume'] * (drink['alcohol_content'] / 100) * 0.8
                drinks_table_data.append([
                    drink['name'],
                    str(drink['volume']),
                    f"{drink['alcohol_content']:.1f}",
                    drink['time'].strftime('%H:%M') if isinstance(drink['time'], datetime) else str(drink['time']),
                    f"{alcohol_grams:.1f}"
                ])
            
            drinks_table = Table(drinks_table_data, colWidths=[4*cm, 2*cm, 2*cm, 2*cm, 2*cm])
            drinks_table.setStyle(TableStyle([
                ('BACKGROUND', (0, 0), (-1, 0), colors.grey),
                ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
                ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
                ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                ('FONTSIZE', (0, 0), (-1, 0), 12),
                ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
                ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
                ('GRID', (0, 0), (-1, -1), 1, colors.black)
            ]))
            story.append(drinks_table)
            story.append(Spacer(1, 20))
        
        self.progress_updated.emit(60)
        
        # Ergebnisse
        if 'results' in self.data and self.data['results']:
            story.append(Paragraph("Berechnungsergebnisse", styles['Heading2']))
            
            results_table_data = [['Modell', 'Aktuelle BAK', 'Max. BAK', 'Zeit bis 0.5‰', 'Zeit bis 0.0‰']]
            
            for model, result in self.data['results'].items():
                results_table_data.append([
                    model,
                    f"{result.get('current_bac', 0):.2f} ‰",
                    f"{result.get('peak_bac', 0):.2f} ‰",
                    result.get('time_to_03', '--').strftime('%H:%M') if result.get('time_to_03') else '--',
                    result.get('time_to_00', '--').strftime('%H:%M') if result.get('time_to_00') else '--'
                ])
            
            results_table = Table(results_table_data, colWidths=[3*cm, 2.5*cm, 2.5*cm, 2.5*cm, 2.5*cm])
            results_table.setStyle(TableStyle([
                ('BACKGROUND', (0, 0), (-1, 0), colors.grey),
                ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
                ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
                ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                ('FONTSIZE', (0, 0), (-1, 0), 12),
                ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
                ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
                ('GRID', (0, 0), (-1, -1), 1, colors.black)
            ]))
            story.append(results_table)
            story.append(Spacer(1, 20))
        
        self.progress_updated.emit(80)
        
        # Chart hinzufügen (falls verfügbar)
        if 'chart_data' in self.data and self.data['chart_data']:
            story.append(Paragraph("BAK-Verlaufsdiagramm", styles['Heading2']))
            
            # Erstelle Chart als Bild
            chart_path = self.file_path.replace('.pdf', '_chart.png')
            self._create_chart_image(chart_path)
            
            if os.path.exists(chart_path):
                chart_img = Image(chart_path, width=15*cm, height=10*cm)
                story.append(chart_img)
                # Temporäres Chart-Bild löschen
                os.remove(chart_path)
        
        # Disclaimer
        story.append(Spacer(1, 30))
        disclaimer_style = ParagraphStyle(
            'Disclaimer',
            parent=styles['Normal'],
            fontSize=10,
            textColor=colors.red,
            alignment=TA_CENTER
        )
        story.append(Paragraph(
            "<b>WICHTIGER HINWEIS:</b> Diese Berechnung dient nur zu Informationszwecken. "
            "Die tatsächliche Blutalkoholkonzentration kann von den berechneten Werten abweichen. "
            "Fahren Sie niemals unter Alkoholeinfluss!",
            disclaimer_style
        ))
        
        # PDF erstellen
        doc.build(story)
        self.progress_updated.emit(100)
    
    def _create_chart_image(self, file_path: str):
        """Erstellt ein Chart-Bild für PDF-Export"""
        try:
            import matplotlib.pyplot as plt
            
            fig, ax = plt.subplots(figsize=(12, 8))
            
            colors = ['#2196F3', '#FF9800', '#4CAF50', '#9C27B0']
            
            for i, (model, result) in enumerate(self.data['chart_data'].items()):
                if 'bac_values' in result and result['bac_values']:
                    times = [point[0] for point in result['bac_values']]
                    bac_values = [point[1] for point in result['bac_values']]
                    
                    color = colors[i % len(colors)]
                    ax.plot(times, bac_values, label=f"{model}", color=color, linewidth=2)
            
            # Grenzwerte
            ax.axhline(y=0.3, color='orange', linestyle='--', alpha=0.7, label='0.3‰')
            ax.axhline(y=0.5, color='red', linestyle='--', alpha=0.7, label='0.5‰')
            ax.axhline(y=1.1, color='darkred', linestyle='--', alpha=0.7, label='1.1‰')
            
            ax.set_xlabel('Zeit')
            ax.set_ylabel('BAK (‰)')
            ax.set_title('Blutalkoholkonzentrations-Verlauf')
            ax.legend()
            ax.grid(True, alpha=0.3)
            
            plt.tight_layout()
            plt.savefig(file_path, dpi=300, bbox_inches='tight')
            plt.close()
            
        except Exception as e:
            print(f"Fehler beim Erstellen des Chart-Bildes: {e}")
    
    def _export_csv(self):
        """Exportiert als CSV"""
        self.progress_updated.emit(20)
        
        with open(self.file_path, 'w', newline='', encoding='utf-8') as csvfile:
            writer = csv.writer(csvfile, delimiter=';')
            
            # Header
            writer.writerow(['BAK-Kalkulator Export'])
            writer.writerow(['Erstellt am:', datetime.now().strftime('%d.%m.%Y %H:%M')])
            writer.writerow([])
            
            self.progress_updated.emit(40)
            
            # Personendaten
            if 'person_data' in self.data:
                writer.writerow(['Personendaten'])
                person = self.data['person_data']
                writer.writerow(['Geschlecht:', person.get('gender', 'N/A')])
                writer.writerow(['Alter:', f"{person.get('age', 'N/A')} Jahre"])
                writer.writerow(['Größe:', f"{person.get('height', 'N/A')} cm"])
                writer.writerow(['Gewicht:', f"{person.get('weight', 'N/A')} kg"])
                writer.writerow(['Körperfettanteil:', f"{person.get('body_fat', 'N/A')}%"])
                writer.writerow(['Trinkgewohnheit:', person.get('drinking_habit', 'N/A')])
                writer.writerow([])
            
            self.progress_updated.emit(60)
            
            # Getränke
            if 'drinks_data' in self.data and self.data['drinks_data']:
                writer.writerow(['Konsumierte Getränke'])
                writer.writerow(['Getränk', 'Menge (ml)', 'Alkohol (%)', 'Zeit', 'Alkohol (g)'])
                
                for drink in self.data['drinks_data']:
                    alcohol_grams = drink['volume'] * (drink['alcohol_content'] / 100) * 0.8
                    writer.writerow([
                        drink['name'],
                        drink['volume'],
                        f"{drink['alcohol_content']:.1f}",
                        drink['time'].strftime('%H:%M') if isinstance(drink['time'], datetime) else str(drink['time']),
                        f"{alcohol_grams:.1f}"
                    ])
                writer.writerow([])
            
            self.progress_updated.emit(80)
            
            # Ergebnisse
            if 'results' in self.data and self.data['results']:
                writer.writerow(['Berechnungsergebnisse'])
                writer.writerow(['Modell', 'Aktuelle BAK', 'Max. BAK', 'Zeit bis 0.5‰', 'Zeit bis 0.0‰'])
                
                for model, result in self.data['results'].items():
                    writer.writerow([
                        model,
                        f"{result.get('current_bac', 0):.2f}",
                        f"{result.get('peak_bac', 0):.2f}",
                        result.get('time_to_03', '--').strftime('%H:%M') if result.get('time_to_03') else '--',
                        result.get('time_to_00', '--').strftime('%H:%M') if result.get('time_to_00') else '--'
                    ])
        
        self.progress_updated.emit(100)
    
    def _export_json(self):
        """Exportiert als JSON"""
        self.progress_updated.emit(20)
        
        # Prepare data for JSON serialization
        json_data = {}
        
        # Convert datetime objects to strings
        if 'drinks_data' in self.data:
            drinks_data = []
            for drink in self.data['drinks_data']:
                drink_copy = drink.copy()
                if isinstance(drink_copy.get('time'), datetime):
                    drink_copy['time'] = drink_copy['time'].isoformat()
                drinks_data.append(drink_copy)
            json_data['drinks_data'] = drinks_data
        
        self.progress_updated.emit(60)
        
        # Convert results datetime objects
        if 'results' in self.data:
            results_data = {}
            for model, result in self.data['results'].items():
                result_copy = result.copy()
                for key, value in result_copy.items():
                    if isinstance(value, datetime):
                        result_copy[key] = value.isoformat()
                results_data[model] = result_copy
            json_data['results'] = results_data
        
        # Copy other data as-is
        for key, value in self.data.items():
            if key not in ['drinks_data', 'results']:
                json_data[key] = value
        
        with open(self.file_path, 'w', encoding='utf-8') as jsonfile:
            json.dump(json_data, jsonfile, indent=2, ensure_ascii=False)
        
        self.progress_updated.emit(100)
    
    def _export_excel(self):
        """Exportiert als Excel"""
        try:
            import openpyxl
            from openpyxl.styles import Font, PatternFill, Alignment
        except ImportError:
            raise ImportError("openpyxl ist nicht installiert. Bitte installieren Sie es mit: pip install openpyxl")
        
        self.progress_updated.emit(20)
        
        workbook = openpyxl.Workbook()
        
        # Personendaten Sheet
        if 'person_data' in self.data:
            ws_person = workbook.active
            ws_person.title = "Personendaten"
            
            person = self.data['person_data']
            ws_person['A1'] = "Personendaten"
            ws_person['A1'].font = Font(bold=True, size=14)
            
            data_rows = [
                ['Geschlecht:', person.get('gender', 'N/A')],
                ['Alter:', f"{person.get('age', 'N/A')} Jahre"],
                ['Größe:', f"{person.get('height', 'N/A')} cm"],
                ['Gewicht:', f"{person.get('weight', 'N/A')} kg"],
                ['Körperfettanteil:', f"{person.get('body_fat', 'N/A')}%"],
                ['Trinkgewohnheit:', person.get('drinking_habit', 'N/A')]
            ]
            
            for i, (label, value) in enumerate(data_rows, start=3):
                ws_person[f'A{i}'] = label
                ws_person[f'B{i}'] = value
                ws_person[f'A{i}'].font = Font(bold=True)
        
        self.progress_updated.emit(50)
        
        # Getränke Sheet
        if 'drinks_data' in self.data and self.data['drinks_data']:
            ws_drinks = workbook.create_sheet("Getränke")
            
            headers = ['Getränk', 'Menge (ml)', 'Alkohol (%)', 'Zeit', 'Alkohol (g)']
            for col, header in enumerate(headers, start=1):
                cell = ws_drinks.cell(row=1, column=col, value=header)
                cell.font = Font(bold=True)
                cell.fill = PatternFill(start_color="CCCCCC", end_color="CCCCCC", fill_type="solid")
            
            for row, drink in enumerate(self.data['drinks_data'], start=2):
                alcohol_grams = drink['volume'] * (drink['alcohol_content'] / 100) * 0.8
                ws_drinks.cell(row=row, column=1, value=drink['name'])
                ws_drinks.cell(row=row, column=2, value=drink['volume'])
                ws_drinks.cell(row=row, column=3, value=drink['alcohol_content'])
                ws_drinks.cell(row=row, column=4, value=drink['time'].strftime('%H:%M') if isinstance(drink['time'], datetime) else str(drink['time']))
                ws_drinks.cell(row=row, column=5, value=round(alcohol_grams, 1))
        
        self.progress_updated.emit(80)
        
        # Ergebnisse Sheet
        if 'results' in self.data and self.data['results']:
            ws_results = workbook.create_sheet("Ergebnisse")
            
            headers = ['Modell', 'Aktuelle BAK', 'Max. BAK', 'Zeit bis 0.5‰', 'Zeit bis 0.0‰']
            for col, header in enumerate(headers, start=1):
                cell = ws_results.cell(row=1, column=col, value=header)
                cell.font = Font(bold=True)
                cell.fill = PatternFill(start_color="CCCCCC", end_color="CCCCCC", fill_type="solid")
            
            for row, (model, result) in enumerate(self.data['results'].items(), start=2):
                ws_results.cell(row=row, column=1, value=model)
                ws_results.cell(row=row, column=2, value=f"{result.get('current_bac', 0):.2f}")
                ws_results.cell(row=row, column=3, value=f"{result.get('peak_bac', 0):.2f}")
                ws_results.cell(row=row, column=4, value=result.get('time_to_03', '--').strftime('%H:%M') if result.get('time_to_03') else '--')
                ws_results.cell(row=row, column=5, value=result.get('time_to_00', '--').strftime('%H:%M') if result.get('time_to_00') else '--')
        
        workbook.save(self.file_path)
        self.progress_updated.emit(100)

class ExportManager(QObject):
    """Manager für Export-Funktionen"""
    
    export_started = pyqtSignal()
    export_progress = pyqtSignal(int)
    export_finished = pyqtSignal(bool, str)  # success, message
    
    def __init__(self, parent: QWidget = None):
        super().__init__()
        self.parent = parent
        self.export_thread = None
    
    def export_to_pdf(self, data: Dict):
        """Exportiert Daten als PDF"""
        file_path, _ = QFileDialog.getSaveFileName(
            self.parent,
            "PDF exportieren",
            f"BAK_Berechnung_{datetime.now().strftime('%Y%m%d_%H%M%S')}.pdf",
            "PDF Dateien (*.pdf)"
        )
        
        if file_path:
            self._start_export('pdf', file_path, data)
    
    def export_to_csv(self, data: Dict):
        """Exportiert Daten als CSV"""
        file_path, _ = QFileDialog.getSaveFileName(
            self.parent,
            "CSV exportieren",
            f"BAK_Berechnung_{datetime.now().strftime('%Y%m%d_%H%M%S')}.csv",
            "CSV Dateien (*.csv)"
        )
        
        if file_path:
            self._start_export('csv', file_path, data)
    
    def export_to_excel(self, data: Dict):
        """Exportiert Daten als Excel"""
        file_path, _ = QFileDialog.getSaveFileName(
            self.parent,
            "Excel exportieren",
            f"BAK_Berechnung_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx",
            "Excel Dateien (*.xlsx)"
        )
        
        if file_path:
            self._start_export('excel', file_path, data)
    
    def export_to_json(self, data: Dict):
        """Exportiert Daten als JSON"""
        file_path, _ = QFileDialog.getSaveFileName(
            self.parent,
            "JSON exportieren",
            f"BAK_Berechnung_{datetime.now().strftime('%Y%m%d_%H%M%S')}.json",
            "JSON Dateien (*.json)"
        )
        
        if file_path:
            self._start_export('json', file_path, data)
    
    def _start_export(self, export_type: str, file_path: str, data: Dict):
        """Startet den Export in einem separaten Thread"""
        if self.export_thread and self.export_thread.isRunning():
            QMessageBox.warning(self.parent, "Export läuft", "Es läuft bereits ein Export. Bitte warten Sie.")
            return
        
        self.export_started.emit()
        
        self.export_thread = ExportThread(export_type, file_path, data)
        self.export_thread.progress_updated.connect(self.export_progress.emit)
        self.export_thread.export_finished.connect(self.export_finished.emit)
        self.export_thread.start() 